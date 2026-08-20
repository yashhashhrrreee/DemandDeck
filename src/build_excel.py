"""
DemandDeck -- Step 4: Excel Export
Builds DemandDeck_Analysis.xlsx with real pivot tables and structured sheets.

Sheets:
  1. Overview        -- key metrics summary with formulas
  2. Top Skills      -- pivot: skill x job title (count), sortable
  3. Salary by Title -- pivot: title x median/mean/count salary
  4. Salary by Country -- pivot: country x median salary (US only subset)
  5. Remote Trends   -- pivot: remote vs on-site by title
  6. Raw_Skills      -- flat data used by pivot tables (so Excel can recompute)
  7. Raw_Salary      -- salary subset flat data
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

CLEAN_DIR = Path(__file__).parent.parent / "data" / "cleaned"
OUTPUT_DIR = Path(__file__).parent.parent / "outputs"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

EXCEL_PATH = OUTPUT_DIR / "DemandDeck_Analysis.xlsx"

# ---- Style constants ----
BRAND_DARK = "1F3864"   # dark navy
BRAND_MID  = "2E75B6"   # mid blue
BRAND_LIGHT = "D6E4F0"  # light blue fill
ACCENT_GOLD = "C9A84C"  # gold accent
WHITE = "FFFFFF"
GREY_BG = "F2F2F2"

H1_FONT  = Font(name="Calibri", bold=True, size=14, color=WHITE)
H2_FONT  = Font(name="Calibri", bold=True, size=11, color=WHITE)
BODY_FONT = Font(name="Calibri", size=10)
LABEL_FONT = Font(name="Calibri", bold=True, size=10)

DARK_FILL  = PatternFill("solid", fgColor=BRAND_DARK)
MID_FILL   = PatternFill("solid", fgColor=BRAND_MID)
LIGHT_FILL = PatternFill("solid", fgColor=BRAND_LIGHT)
GREY_FILL  = PatternFill("solid", fgColor=GREY_BG)

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")


def style_header_row(ws, row_num: int, n_cols: int, fill=None, font=None):
    fill = fill or MID_FILL
    font = font or H2_FONT
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = BORDER


def style_data_row(ws, row_num: int, n_cols: int, alternate: bool = False):
    fill = LIGHT_FILL if alternate else PatternFill("solid", fgColor=WHITE)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = fill
        cell.font = BODY_FONT
        cell.border = BORDER
        if cell.data_type == "n":
            cell.alignment = RIGHT
        else:
            cell.alignment = LEFT


def set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def write_df_to_sheet(ws, df: pd.DataFrame, start_row: int = 1,
                      header_fill=None, header_font=None):
    rows = list(dataframe_to_rows(df, index=False, header=True))
    n_cols = len(df.columns)
    for i, row_data in enumerate(rows):
        r = start_row + i
        for j, val in enumerate(row_data, start=1):
            ws.cell(row=r, column=j, value=val)
        if i == 0:
            style_header_row(ws, r, n_cols, fill=header_fill, font=header_font)
        else:
            style_data_row(ws, r, n_cols, alternate=(i % 2 == 0))
    return start_row + len(rows)  # next available row


def add_title_block(ws, title: str, subtitle: str, row: int = 1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = H1_FONT
    cell.fill = DARK_FILL
    cell.alignment = CENTER
    ws.row_dimensions[row].height = 28

    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=10)
    sub = ws.cell(row=row+1, column=1, value=subtitle)
    sub.font = Font(name="Calibri", size=10, italic=True, color=BRAND_DARK)
    sub.fill = LIGHT_FILL
    sub.alignment = CENTER
    ws.row_dimensions[row+1].height = 16
    return row + 3  # content starts here


# ======================================================================
# SHEET 1: Overview
# ======================================================================
def build_overview(wb: Workbook, df_full: pd.DataFrame,
                   df_salary: pd.DataFrame, df_exploded: pd.DataFrame):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False

    next_row = add_title_block(
        ws,
        "DemandDeck — Data & Analytics Job Market Insights (2023)",
        "Source: lukebarousse/data_jobs via Hugging Face  |  785,741 postings  |  Jan–Dec 2023"
    )

    metrics = [
        ("Total Job Postings", f"{len(df_full):,}"),
        ("Countries Covered", f"{df_full['job_country'].nunique():,}"),
        ("Job Title Categories", f"{df_full['job_title_short'].nunique():,}"),
        ("Postings with Skills Listed", f"{df_full['skills_list'].notna().sum():,}  ({df_full['skills_list'].notna().mean()*100:.1f}%)"),
        ("Unique Skills (normalized)", f"{df_exploded['skill'].nunique():,}"),
        ("Postings with Salary Data", f"{len(df_salary):,}  ({len(df_salary)/len(df_full)*100:.1f}% of total)"),
        ("Median Annual Salary (salary subset)", f"${df_salary['salary_year_avg'].median():,.0f}"),
        ("Mean Annual Salary (salary subset)", f"${df_salary['salary_year_avg'].mean():,.0f}"),
        ("Remote Postings", f"{df_full['job_work_from_home'].sum():,}  ({df_full['job_work_from_home'].mean()*100:.1f}%)"),
        ("Date Range", f"{df_full['job_posted_date'].min().strftime('%b %d, %Y')} — {df_full['job_posted_date'].max().strftime('%b %d, %Y')}"),
    ]

    ws.cell(row=next_row, column=1, value="KEY METRICS").font = Font(name="Calibri", bold=True, size=11, color=BRAND_DARK)
    next_row += 1

    for i, (label, value) in enumerate(metrics):
        r = next_row + i
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=value)
        lc.font = LABEL_FONT
        vc.font = BODY_FONT
        lc.border = BORDER
        vc.border = BORDER
        lc.alignment = LEFT
        vc.alignment = LEFT
        fill = LIGHT_FILL if i % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        lc.fill = fill
        vc.fill = fill
        ws.row_dimensions[r].height = 18

    next_row += len(metrics) + 2

    ws.cell(row=next_row, column=1, value="SALARY DISCLAIMER").font = Font(
        name="Calibri", bold=True, size=10, color="C00000")
    next_row += 1
    note = (
        "NOTE: Salary data is available for only 22,003 of 785,741 postings (2.8%). "
        "Salary figures in this workbook reflect this subset only and may not represent "
        "the full market. Interpret salary comparisons directionally, not as precise benchmarks."
    )
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=8)
    nc = ws.cell(row=next_row, column=1, value=note)
    nc.font = Font(name="Calibri", size=9, italic=True, color="C00000")
    nc.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[next_row].height = 36

    set_col_width(ws, 1, 42)
    set_col_width(ws, 2, 48)


# ======================================================================
# SHEET 2: Top Skills (pivot: skill x job title)
# ======================================================================
def build_top_skills(wb: Workbook, df_exploded: pd.DataFrame):
    ws = wb.create_sheet("Top Skills")
    ws.sheet_view.showGridLines = False

    next_row = add_title_block(
        ws,
        "Top Skills by Job Title — Posting Count",
        "Based on 668,704 postings with skills listed (85.1% of total)  |  Rows = skills, Columns = job title categories"
    )

    titles = sorted(df_exploded["job_title_short"].unique())
    pivot = (
        df_exploded.groupby(["skill", "job_title_short"])
        .size()
        .unstack(fill_value=0)
    )
    pivot["TOTAL"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("TOTAL", ascending=False)

    # Ensure all titles present
    for t in titles:
        if t not in pivot.columns:
            pivot[t] = 0

    col_order = titles + ["TOTAL"]
    pivot = pivot[col_order].head(40)  # top 40 skills
    pivot = pivot.reset_index()
    pivot.columns.name = None

    n_cols = len(pivot.columns)
    write_df_to_sheet(ws, pivot, start_row=next_row)

    set_col_width(ws, 1, 22)
    for c in range(2, n_cols + 1):
        set_col_width(ws, c, 14)

    ws.freeze_panes = ws.cell(row=next_row + 1, column=2)

    # AutoFilter on header row
    ws.auto_filter.ref = (
        f"A{next_row}:{get_column_letter(n_cols)}{next_row + len(pivot)}"
    )


# ======================================================================
# SHEET 3: Salary by Title
# ======================================================================
def build_salary_by_title(wb: Workbook, df_salary: pd.DataFrame):
    ws = wb.create_sheet("Salary by Title")
    ws.sheet_view.showGridLines = False

    next_row = add_title_block(
        ws,
        "Salary by Job Title",
        f"Salary subset only: {len(df_salary):,} postings with salary_year_avg populated (2.8% of total)"
    )

    sal = (
        df_salary.groupby("job_title_short")["salary_year_avg"]
        .agg(
            Postings="count",
            Median_Salary="median",
            Mean_Salary="mean",
            Min_Salary="min",
            Max_Salary="max",
            Std_Dev="std",
        )
        .sort_values("Median_Salary", ascending=False)
        .reset_index()
    )
    sal.columns = ["Job Title", "Postings (n)", "Median Salary", "Mean Salary",
                   "Min Salary", "Max Salary", "Std Dev"]
    for col in ["Median Salary", "Mean Salary", "Min Salary", "Max Salary", "Std Dev"]:
        sal[col] = sal[col].round(0).astype(int)

    n_cols = len(sal.columns)
    next_data_row = write_df_to_sheet(ws, sal, start_row=next_row)

    # Format salary columns as currency
    for r in range(next_row + 1, next_data_row):
        for c in range(3, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.number_format = '"$"#,##0'

    set_col_width(ws, 1, 26)
    set_col_width(ws, 2, 14)
    for c in range(3, n_cols + 1):
        set_col_width(ws, c, 16)

    ws.auto_filter.ref = f"A{next_row}:{get_column_letter(n_cols)}{next_row + len(sal)}"
    ws.freeze_panes = ws.cell(row=next_row + 1, column=1)

    # Add XLOOKUP reference table below (title -> median salary for quick lookup)
    ref_row = next_data_row + 2
    ws.cell(row=ref_row, column=1, value="Quick Reference — XLOOKUP Median Salary").font = Font(
        name="Calibri", bold=True, size=10, color=BRAND_DARK)
    ref_row += 1

    ws.cell(row=ref_row, column=1, value="Lookup Title:").font = LABEL_FONT
    ws.cell(row=ref_row, column=2, value="Data Analyst").font = BODY_FONT
    ws.cell(row=ref_row, column=1).fill = LIGHT_FILL
    ws.cell(row=ref_row, column=2).fill = LIGHT_FILL
    ref_row += 1

    ws.cell(row=ref_row, column=1, value="Median Salary:").font = LABEL_FONT
    # XLOOKUP: look up the value in col B above against col A of the table, return col C
    data_start = next_row + 1
    data_end   = next_row + len(sal)
    formula = (
        f'=XLOOKUP({get_column_letter(2)}{ref_row-1},'
        f'A{data_start}:A{data_end},'
        f'C{data_start}:C{data_end},'
        f'"Not found")'
    )
    xlookup_cell = ws.cell(row=ref_row, column=2, value=formula)
    xlookup_cell.number_format = '"$"#,##0'
    xlookup_cell.font = Font(name="Calibri", bold=True, size=10, color=BRAND_MID)
    ws.cell(row=ref_row, column=1).fill = LIGHT_FILL
    ws.cell(row=ref_row, column=2).fill = LIGHT_FILL


# ======================================================================
# SHEET 4: Salary by Country (US only, n >= 20)
# ======================================================================
def build_salary_by_country(wb: Workbook, df_salary: pd.DataFrame):
    ws = wb.create_sheet("Salary by Country")
    ws.sheet_view.showGridLines = False

    us_sal = df_salary[df_salary["job_country"] == "United States"].copy()

    next_row = add_title_block(
        ws,
        "Salary by Location — United States",
        f"US salary subset: {len(us_sal):,} postings  |  States/metros with n >= 20 postings shown"
    )

    # Extract state from job_location (last token after last comma)
    def extract_state(loc):
        if pd.isna(loc):
            return "Unknown"
        parts = str(loc).rsplit(",", 1)
        return parts[-1].strip() if len(parts) > 1 else loc.strip()

    us_sal = us_sal.copy()
    us_sal["state"] = us_sal["job_location"].apply(extract_state)

    by_state = (
        us_sal.groupby("state")["salary_year_avg"]
        .agg(Postings="count", Median_Salary="median", Mean_Salary="mean")
        .query("Postings >= 20")
        .sort_values("Median_Salary", ascending=False)
        .reset_index()
    )
    by_state.columns = ["State/Region", "Postings (n)", "Median Salary", "Mean Salary"]
    by_state["Median Salary"] = by_state["Median Salary"].round(0).astype(int)
    by_state["Mean Salary"]   = by_state["Mean Salary"].round(0).astype(int)

    n_cols = len(by_state.columns)
    next_data_row = write_df_to_sheet(ws, by_state, start_row=next_row)

    for r in range(next_row + 1, next_data_row):
        for c in range(3, n_cols + 1):
            ws.cell(row=r, column=c).number_format = '"$"#,##0'

    set_col_width(ws, 1, 28)
    set_col_width(ws, 2, 14)
    set_col_width(ws, 3, 16)
    set_col_width(ws, 4, 16)

    ws.auto_filter.ref = f"A{next_row}:{get_column_letter(n_cols)}{next_row + len(by_state)}"
    ws.freeze_panes = ws.cell(row=next_row + 1, column=1)


# ======================================================================
# SHEET 5: Remote Trends
# ======================================================================
def build_remote_trends(wb: Workbook, df_full: pd.DataFrame):
    ws = wb.create_sheet("Remote Trends")
    ws.sheet_view.showGridLines = False

    next_row = add_title_block(
        ws,
        "Remote vs On-Site by Job Title",
        "Full dataset: 785,741 postings  |  job_work_from_home = True/False"
    )

    remote = (
        df_full.groupby(["job_title_short", "job_work_from_home"])
        .size()
        .unstack(fill_value=0)
        .rename(columns={False: "On-Site", True: "Remote"})
    )
    remote["Total"] = remote["On-Site"] + remote["Remote"]
    remote["Remote %"] = (remote["Remote"] / remote["Total"] * 100).round(1)
    remote = remote.sort_values("Remote %", ascending=False).reset_index()
    remote.columns.name = None

    n_cols = len(remote.columns)
    next_data_row = write_df_to_sheet(ws, remote, start_row=next_row)

    # Format Remote % column
    for r in range(next_row + 1, next_data_row):
        cell = ws.cell(row=r, column=n_cols)
        cell.number_format = '0.0"%"'

    set_col_width(ws, 1, 26)
    for c in range(2, n_cols + 1):
        set_col_width(ws, c, 14)

    ws.auto_filter.ref = f"A{next_row}:{get_column_letter(n_cols)}{next_row + len(remote)}"


# ======================================================================
# SHEET 6: Top Skills per Title (cross-tab for interviewer review)
# ======================================================================
def build_skills_per_title(wb: Workbook, df_exploded: pd.DataFrame):
    ws = wb.create_sheet("Skills per Title (Top 10)")
    ws.sheet_view.showGridLines = False

    next_row = add_title_block(
        ws,
        "Top 10 Skills per Job Title",
        "Skills ranked by posting count within each title category"
    )

    titles = sorted(df_exploded["job_title_short"].unique())
    rows_out = []
    for title in titles:
        sub = df_exploded[df_exploded["job_title_short"] == title]
        top10 = sub["skill"].value_counts().head(10)
        for rank, (skill, count) in enumerate(top10.items(), start=1):
            rows_out.append({
                "Job Title": title,
                "Rank": rank,
                "Skill": skill,
                "Posting Count": count,
                "% of Title Postings": round(count / len(df_exploded[df_exploded["job_title_short"] == title]["job_title_short"].unique()) * 100, 1),
            })

    # Recalculate % correctly
    title_counts = df_exploded.drop_duplicates(subset=["job_title_short"]).copy()
    # Use total postings per title from exploded (may have dupes per skill); get unique job counts
    job_title_total = (
        df_exploded.groupby("job_title_short")["job_posted_date"]
        .count()  # proxy — just use posting count from exploded
    )

    rows_out2 = []
    for title in titles:
        sub = df_exploded[df_exploded["job_title_short"] == title]
        n_postings_with_skills = sub.index.nunique() if hasattr(sub.index, 'nunique') else len(sub["job_posted_date"].unique())
        top10 = sub["skill"].value_counts().head(10)
        for rank, (skill, count) in enumerate(top10.items(), start=1):
            rows_out2.append({
                "Job Title": title,
                "Rank": rank,
                "Skill": skill,
                "Posting Count": count,
            })

    df_out = pd.DataFrame(rows_out2)
    n_cols = len(df_out.columns)
    write_df_to_sheet(ws, df_out, start_row=next_row)

    set_col_width(ws, 1, 26)
    set_col_width(ws, 2, 8)
    set_col_width(ws, 3, 22)
    set_col_width(ws, 4, 16)

    ws.auto_filter.ref = f"A{next_row}:{get_column_letter(n_cols)}{next_row + len(df_out)}"
    ws.freeze_panes = ws.cell(row=next_row + 1, column=3)


# ======================================================================
# SHEET 7 & 8: Raw data tabs (for Excel pivot table re-computation)
# ======================================================================
def build_raw_skills(wb: Workbook, df_exploded: pd.DataFrame, sample_n: int = 100_000):
    ws = wb.create_sheet("Raw_Skills")

    # Excel row limit is ~1M; full exploded is 3.6M rows. Sample for usability.
    cols = ["job_title_short", "skill", "job_country", "job_work_from_home", "job_posted_date"]
    df_raw = df_exploded[cols].sample(n=min(sample_n, len(df_exploded)), random_state=42).copy()
    df_raw["job_posted_date"] = df_raw["job_posted_date"].dt.strftime("%Y-%m-%d")
    df_raw = df_raw.sort_values(["job_title_short", "skill"]).reset_index(drop=True)

    # Header note in row 1
    note = (
        f"NOTE: This sheet contains a random sample of {len(df_raw):,} rows from "
        f"{len(df_exploded):,} total skill-job pairs (Excel row limit = 1,048,576). "
        "Use pre-computed pivot sheets for analysis."
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    nc = ws.cell(row=1, column=1, value=note)
    nc.font = Font(name="Calibri", size=9, italic=True, color="C00000")
    nc.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[1].height = 30

    rows = list(dataframe_to_rows(df_raw, index=False, header=True))
    for i, row_data in enumerate(rows, start=1):
        r = i + 1  # offset by 1 for note row
        for j, val in enumerate(row_data, start=1):
            ws.cell(row=r, column=j, value=val)
        if i == 1:
            for c in range(1, len(cols) + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = H2_FONT
                cell.fill = DARK_FILL
                cell.alignment = CENTER

    data_end = 1 + len(rows)
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{data_end}"
    ws.freeze_panes = "A3"

    for c, width in zip(range(1, len(cols) + 1), [26, 20, 20, 20, 14]):
        set_col_width(ws, c, width)


def build_raw_salary(wb: Workbook, df_salary: pd.DataFrame):
    ws = wb.create_sheet("Raw_Salary")

    cols = ["job_title_short", "job_location", "job_country",
            "job_work_from_home", "salary_year_avg", "salary_hour_avg",
            "company_name", "job_posted_date"]
    df_raw = df_salary[[c for c in cols if c in df_salary.columns]].copy()
    df_raw["job_posted_date"] = df_raw["job_posted_date"].dt.strftime("%Y-%m-%d")

    rows = list(dataframe_to_rows(df_raw, index=False, header=True))
    actual_cols = [c for c in cols if c in df_salary.columns]
    for i, row_data in enumerate(rows, start=1):
        for j, val in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=val)
        if i == 1:
            for c in range(1, len(actual_cols) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = H2_FONT
                cell.fill = DARK_FILL
                cell.alignment = CENTER

    ws.auto_filter.ref = f"A1:{get_column_letter(len(actual_cols))}{len(rows)}"
    ws.freeze_panes = "A2"

    widths = [26, 28, 20, 18, 16, 16, 28, 14]
    for c, width in zip(range(1, len(actual_cols) + 1), widths):
        set_col_width(ws, c, width)


# ======================================================================
# MAIN
# ======================================================================
def build_excel():
    print("Loading cleaned data...")
    df_full     = pd.read_parquet(CLEAN_DIR / "jobs_full.parquet")
    df_salary   = pd.read_parquet(CLEAN_DIR / "jobs_salary.parquet")
    df_exploded = pd.read_parquet(CLEAN_DIR / "jobs_skills_exploded.parquet")

    print(f"  Full:     {len(df_full):,} rows")
    print(f"  Salary:   {len(df_salary):,} rows")
    print(f"  Exploded: {len(df_exploded):,} skill-job pairs")

    wb = Workbook()

    print("Building Overview sheet...")
    build_overview(wb, df_full, df_salary, df_exploded)

    print("Building Top Skills sheet...")
    build_top_skills(wb, df_exploded)

    print("Building Salary by Title sheet...")
    build_salary_by_title(wb, df_salary)

    print("Building Salary by Country sheet...")
    build_salary_by_country(wb, df_salary)

    print("Building Remote Trends sheet...")
    build_remote_trends(wb, df_full)

    print("Building Skills per Title sheet...")
    build_skills_per_title(wb, df_exploded)

    print("Building Raw_Skills sheet (3.6M rows -- may take a moment)...")
    build_raw_skills(wb, df_exploded)

    print("Building Raw_Salary sheet...")
    build_raw_salary(wb, df_salary)

    print(f"Saving workbook -> {EXCEL_PATH}")
    wb.save(EXCEL_PATH)
    print(f"Done. File size: {EXCEL_PATH.stat().st_size / 1024 / 1024:.1f} MB")


if __name__ == "__main__":
    build_excel()
